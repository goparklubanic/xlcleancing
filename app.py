import shutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
# import formulas
import os

# set resources path
resources_path = os.path.join(os.path.dirname(__file__), 'resources')
processed_path = os.path.join(os.path.dirname(__file__), 'kelar')

# copy template
def copytemplate():
    results_path = os.path.join(os.path.dirname(__file__), 'results')
    template_path = os.path.join(results_path, 'template.xlsx')
    itemlist_path = os.path.join(results_path, 'itemlist.xlsx')
    # copy template to itemlist
    shutil.copyfile(template_path, itemlist_path)
    return itemlist_path

# find excel files in resources folder
def findexcels():
    excel_files = []
    for file in os.listdir(resources_path):
        if file.endswith(('.xlsx', '.xls')) and not file.startswith('~'):
            excel_files.append(os.path.join(resources_path, file))
    return excel_files

# read excel files and extract data
'''
The header is row number 15.
The last rows identified by value "TOTAL" in Column A.
If cell in Column A is empty, then fill it with number 123
Delete row if column A is "PO#, SUBTOTAL, Buyer Dia, Mountings
Put the rest into a dataframe
'''
def readexcels(file):
    # excel_files = findexcels()
    # for file in excel_files:
    print(f"Reading {file}")
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    buyer = ws['B8'].value
    if buyer is None:
        buyer = 'Veronique Oro'
          
    data = []
    header = [cell.value for cell in ws[15]]
    evalheader(header)
    for row in ws.iter_rows(min_row=16, max_row=ws.max_row):
        if row[0].value == 'TOTAL':
            break
        if not row[0].value:
            row[0].value = 123
        if row[0].value in ['PO#', 'SUBTOTAL', 'Buyer Dia', 'Mountings','Mounting','']:
            continue
        data.append([cell.value for cell in row])
        # data[-1].insert(0, buyer)
    df = pd.DataFrame(data, columns=header)
    
    # give header to df
    df.columns = header

    # modify data frame. Put buyer as first header. the value is buyer
    df.insert(0, 'buyer', buyer)
    # print(df)
    return df

def evalheader(header):
    # loop through header, replace all whitespace with empty string
    ui=0
    for i in range(len(header)):
        # if header[i] is none or empty, replace with unknown
        if not header[i]:
            header[i] = 'unknown' + str(ui)
            ui += 1
        header[i] = header[i].replace('\'', '')
        # smallcase header[i]
        header[i] = header[i].lower()
        # replace manufacturing to maklon
        if header[i] == 'manufacturing':
            header[i] = 'maklon'
        # replace space with underscore
        header[i] = header[i].replace(' ', '_')
        # remove dot
        header[i] = header[i].replace('.', '')
        # replace carriage return with underscore
        header[i] = header[i].replace('\n', '_')
        # print(header[i])
    # print(header)
    return header

# filter columns
'''
Only use column PO#, Item No, Metal, Q'ty, Total w't, maklon, Non US Dia, and total.
if maklon is not exist but manufacturing is exist, then use manufacturing to replace 
'''
def colfilter(df):
    df = df[['buyer','po#', 'item_no', 'metal', 'qty', 'total_wt', 'maklon', 'non_us_dia', 'total']]
    return df

# adjust value of column PO#
'''
if PO# value length more than 6, than keep it as temporary variable called ponum
if next PO# value is empty or number, replace it with ponum
'''
def poadjust(df):
    ponum = ''
    for i in range(len(df)):
        po_val = df.loc[i, 'po#']
        # check po# is started with string a-z
        if str(po_val)[0].isalpha():
            # print(f"PO# {df.iloc[i]['po#']} is string")
            ponum = po_val
        else:
            # replace with ponum
            # df.iloc[i]['po#'] = ponum
            df.loc[i,'po#'] = ponum
        # print(f"PO# {df.loc[i,'po#']}")
    return df


# add dataframe to results/itemlist.xlsx
'''
- copy result/template.xlsx to result/itemlist.xlsx
- open result/itemlist.xlsx
- get the last row of data, check if po# is same with new dataframe index 0 po#
- loop the dataframe, append to result/itemlist.xlsx start from last row
'''
def addtolist(df, itemlist_path):
    # open itemlist
    wb = openpyxl.load_workbook(itemlist_path)
    ws = wb.active
    # get the last row of data
    lastrow = ws.max_row+1

    # loop the dataframe, append to result/itemlist.xlsx start from last row
    for i in range(len(df)):
        ws.cell(row=lastrow+i, column=1).value = df.iloc[i]['buyer']
        ws.cell(row=lastrow+i, column=2).value = df.iloc[i]['po#']
        ws.cell(row=lastrow+i, column=3).value = df.iloc[i]['item_no']
        ws.cell(row=lastrow+i, column=4).value = df.iloc[i]['metal']
        ws.cell(row=lastrow+i, column=5).value = df.iloc[i]['qty']
        ws.cell(row=lastrow+i, column=6).value = df.iloc[i]['total_wt']
        ws.cell(row=lastrow+i, column=7).value = df.iloc[i]['maklon']
        ws.cell(row=lastrow+i, column=8).value = df.iloc[i]['non_us_dia']
        ws.cell(row=lastrow+i, column=9).value = df.iloc[i]['total']
    wb.save(itemlist_path)

def resultcleansing():
    # open itemlist
    wb = openpyxl.load_workbook('results/itemlist.xlsx')
    ws = wb.active

    # evaluate column C. delete row if value is empty
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if not row[2].value:
            ws.delete_rows(row[0].row)
    wb.save('results/itemlist.xlsx')



# main function
def main():
    # copy template to itemlist
    itemlist_path = copytemplate()
    excel_files = findexcels()
    for file in excel_files:
        df0 = readexcels(file)
        df1 = colfilter(df0)
        # poadjust(df1)
        df2 = poadjust(df1)
        # print(df2)
        addtolist(df2, itemlist_path)
        # move file to ../kelar/
        os.rename(file, os.path.join(processed_path, os.path.basename(file)))
        resultcleansing()   

if __name__ == '__main__':
    main()